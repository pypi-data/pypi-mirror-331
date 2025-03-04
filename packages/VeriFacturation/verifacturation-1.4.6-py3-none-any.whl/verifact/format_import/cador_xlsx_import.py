from openpyxl import load_workbook
from typing import Dict, Any
import datetime
import polars as pl
from .base_import import BaseImport
from verifact.error import run_error

class CadorXlsxImport(BaseImport):
    def name(self):
        return "CADOR/ComptabilitéExpert (.xlsx)"
    
    def validate_format(self):
        if self.path.suffix.lower() != ".xlsx":
            run_error("Ce format CADOR nécessite un fichier .xlsx")
            return False
        return True

    def process_file(self):
        wb = load_workbook(filename=self.filename)
        ws = wb.worksheets[0]
        
        # Vérifie que le type de fichier est correct
        if ws.cell(1, 1).value != 'Edition Journaux':
            run_error("Il ne s'agit pas d'un journal de vente")
            return []
        
        nb_lignes = ws.max_row
        liste_factures = []
        date_journal = None
        
        # Récupère la liste des factures de ventes
        for ligne in range(1, nb_lignes + 1):
            if isinstance(ws.cell(ligne, 5).value, datetime.datetime):
                date_journal = ws.cell(ligne, 5).value
            
            try:
                facture = {
                    "PieceRef": str(ws.cell(ligne, 2).value),
                    "EcritureDate": self.extract_date(ws.cell(ligne, 1).value, date_journal),
                    "EcritureLib": str(ws.cell(ligne, 3).value),
                    "CompteNum": str(ws.cell(ligne, 4).value),
                    "Debit": float(str(ws.cell(ligne, 6).value)),
                    "Credit": float(str(ws.cell(ligne, 7).value))
                }
            except Exception:
                continue
            
            # Ajoute la facture à la liste si celle-ci est valide
            if self.validate_account(facture):
                liste_factures.append(facture.copy())
        
        # Transforme la liste de factures en dataframe
        self._sales = pl.DataFrame(liste_factures, orient="row")

    def extract_date(
        self, 
        jour_facture: str, 
        date_journal: datetime.datetime) -> datetime.date:
        """Renvoie la date d'une écriture d'un journal Cador"""
        jour = int(jour_facture)
        mois = int(date_journal.month)
        annee = int(date_journal.year)
        return datetime.date(annee, mois, jour)
    
    def validate_account(self, facture: Dict[str, Any]) -> bool:
        """Vérifie si la facture est valide"""
        return (isinstance(facture.get("CompteNum"), str) and 
                isinstance(facture.get("Debit"), (int, float)) and 
                isinstance(facture.get("Credit"), (int, float)))
