""" Baseline rules for checking contents of Excel files. """
import openpyxl
import pandas as pd

from .ten8t_exception import Ten8tException
from .ten8t_format import BM
from .ten8t_result import TR
from .ten8t_util import StrOrNone, str_to_bool

SHEET1 = "Sheet1"
AUTO = "auto"
DESC_DEFAULT = ""
START_ROW_DEFAULT = "1"
VAL_COL_DEFAULT = "B"
DESCRIPTION_COLUMN = "A"


def _column_to_number(column: str) -> int:
    number = 0
    for i, char in enumerate(reversed(column)):
        number += (ord(char.upper()) - 64) * (26 ** i)
    return number


def _get_sheet(wb, sheet_name=None):
    """Ensure a valid sheet is selected based on sheet_name parameter"""
    if sheet_name is None:
        return wb["Sheet1"]
    if sheet_name in wb.sheetnames:
        return wb[sheet_name]
    if len(wb.sheetnames) == 1:
        return wb[wb.sheetnames[0]]
    if "Sheet1" in wb.sheetnames:
        return wb["Sheet1"]

    raise Ten8tException('A sheet name was not specified and sheet1 could not be found.')


def _ensure_row_params(row_end, row_start: int):
    """Ensure the start and end rows parameters are consistent"""
    auto = False

    if row_end is None:
        return row_start, row_start, auto

    if isinstance(row_end, str):
        if row_end.isdigit() and int(row_end) < row_start:
            raise Ten8tException(
                f'Value for end row must be larger than start row {row_start=} {row_end=}')
        if row_end.lower() == AUTO:
            auto = True
            row_end = 1000
    try:
        row_end = int(row_end)
    except ValueError as vex:
        raise Ten8tException("row_end was not a valid integer value") from vex
    return row_start, row_end, auto


def rule_xlsx_a1_pass_fail(wb: openpyxl.workbook.Workbook,
                           sheet_name: StrOrNone = None,
                           desc_col='A',
                           val_col='B',
                           row_start='1',
                           row_end=None):
    """ This is a very blunt instrument that pulls a true/false value out of
        a specific sheet/row/col of an Excel workbook.  It is very unforgiving
        to format changes in the work book

        Using start and end row numbers you can iterate over many items in the notebook.
        If row end is set to 'auto' it will run until the first blank is detected in
        the value column.

        """
    sheet = _get_sheet(wb, sheet_name)

    # Handle Nones.  Presumably this should not be required
    row_start = row_start or '1'
    val_col = val_col or 'B'
    row_start, row_end, auto = _ensure_row_params(row_end, int(row_start))
    val_col = _column_to_number(val_col)

    if desc_col is not None:
        desc_col = _column_to_number(desc_col)

    for row in range(row_start, row_end + 1):

        value = sheet.cell(row=row, column=val_col).value
        if value is None and auto:
            break

        if value is None:
            raise Ten8tException(f'Expected boolean value in row {row}')

        if desc_col is not None:
            desc = sheet.cell(row=row, column=desc_col).value
        else:
            # It is possible not to have a description column
            desc = ""

        if str_to_bool(value):
            yield TR(status=True, msg=f"{BM.expected(desc)}-Passed")
        else:
            yield TR(status=False, msg=f"{BM.expected(desc)}-Failed")


def rule_xlsx_df_pass_fail(df: pd.DataFrame, desc_col: str, val_col: str, skip_on_none=False):
    """
    One could argue this is really a dataframe tool, but it is assumed that the end
    user will load an Excel spreadsheet into a dataframe and then process it.  This
    guy looks at two columns assuming the first row is the column header
    """

    for row in df.values:
        # Make dictionaries because they are easier to look at.
        row_dict = dict(zip(df.columns, row))

        if pd.isnull(row_dict[val_col]):
            if skip_on_none:
                yield TR(status=None, skipped=True,
                         msg=f"Null value detected in column={BM.expected(val_col)}")
            else:
                yield TR(status=False,
                         msg=f"Null value detected in column={BM.expected(val_col)}")
            continue

        if pd.isnull(row_dict[desc_col]):
            if skip_on_none:
                yield TR(status=None, skipped=True,
                         msg=f"Null description detected in column={BM.expected(desc_col)}")
            else:
                yield TR(status=False,
                         msg=f"Null description detected in column={BM.expected(desc_col)}")
            continue

        description = row_dict[desc_col]

        # Very lenient boolean values
        status = str_to_bool(row_dict[val_col])
        if status:
            yield TR(status=True, msg=f"{BM.code(description)}-Passed")
        else:
            yield TR(status=False, msg=f"{BM.code(description)}-Failed")
