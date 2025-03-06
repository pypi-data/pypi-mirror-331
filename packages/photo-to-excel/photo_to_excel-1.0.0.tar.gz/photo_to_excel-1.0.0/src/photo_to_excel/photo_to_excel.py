import os
from PIL import Image
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from photo_to_excel.helpers import _remove_alpha, _rgb_str_format, _image_size_ratio

class PhotoToExcel:

    def __init__(self, img_path:str, pixel_size:list[int]=(64, 64), save_pixelated:bool=False):
        """
        Initialize the class with the path to the image file and the desired resolution of the pixelated image

        Parameters:
        ----------
        img_path : `str`
            Path to the image file
        pixel_size : `list[int]`
            Desired resolution of the image in pixels, default is (64,64)
        """
        self.img_path = img_path
        self.save_directory = os.path.dirname(self.img_path)
        self.img_file_name = os.path.basename(self.img_path).split(".")[0]
        self.pixel_size = pixel_size
        self.save_pixelated = save_pixelated

    @property
    def pixels_rbg(self)->list[list[int]]:
        """
        Change resolution of the image to desired number of pixels (default 64x64)

        Parameters:
        ----------
        save_pixelated : `bool` (default=`False`)
            Save the pixelated image as a new file

        Returns:
        -------
        rgb_list_2d : `list[list[int]]`
            2D list of RGB values of the pixelated image
        """

        # open image (check if path exists)
        try:
            img = Image.open(self.img_path)
        except(FileNotFoundError):
            raise FileNotFoundError("File not found")

        # resize image (to dither in the future?)
        try:
            img_pixelated = img.resize(self.pixel_size)
        except(TypeError, ValueError):
            print(f"ERROR Invalid pixel size")
            raise
        
        # Scale back up using NEAREST to original size
        if self.save_pixelated==True:
            result = img_pixelated.resize(img.size,Image.NEAREST)
            saved_filepath = f"{self.save_directory}/{self.img_file_name}_pixelated.png"
            print(f"saved in {saved_filepath}")
            result.save(saved_filepath)
        elif self.save_pixelated==False:
            pass

        pix_val = list(img_pixelated.getdata())
        self.rgb_list_2d = [pix_val[i:i+self.pixel_size[0]] for i in range(0, len(pix_val), self.pixel_size[1])]

        return self.rgb_list_2d

    def rgb_to_excel(self)->None:  
        """
        Export the pixelated image to an Excel file. This function takes a 2D list of RGB values from the photo (i.e. the output of `pixelate_photo`) and uses them to colour in the cells in Excel.

        Parameters:
        ----------
        rgb_list_2d : `list[list[int]]`
            2D list of RGB values of the pixelated image
        """
        # Create a new workbook and select the active sheet
        workbook = Workbook()
        sheet = workbook.active

        num_rows, num_columns = self.pixel_size[0], self.pixel_size[1]

        # Set the column width and row height to make square cells
        for row_index in range(1, num_rows + 1):
            sheet.row_dimensions[row_index].height = 75/4

        for column_index in range(1, num_columns + 1):
            col_letter = get_column_letter(column_index)
            sheet.column_dimensions[col_letter].width = 12.43/4
        
        for row_index, row in enumerate(self.pixels_rbg, start=1):
            for col_index, rgb_color in enumerate(row, start=1):
                # Convert RGB to Color object
                rgb_color = _remove_alpha(rgb_color)
                color = PatternFill(start_color=_rgb_str_format(rgb_color), end_color=_rgb_str_format(rgb_color), fill_type="solid")

                # Get the column letter
                col_letter = get_column_letter(col_index)

                # Apply the color to the cell
                sheet[f"{col_letter}{row_index}"].fill = color
                # Debugging: Print RGB values for 'AD1'
                # if col_index == 30 and row_index == 1:
                #     print(f"RGB for 'AD1': {rgb_color}")
    
        workbook.save(f"{self.save_directory}/{self.img_file_name}.xlsx")

