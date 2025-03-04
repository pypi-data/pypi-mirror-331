import csv
import io
import logging
import os
from typing import Optional, Union

import duckdb
import numpy as np
import pandas as pd
import prqlc
import sqlalchemy as sa
from openpyxl import load_workbook
from pdfminer.high_level import LAParams, extract_pages
from pdfminer.layout import LTChar, LTTextBox

import els.config as ec
import els.core as el
import els.xl as xl


def push_frame(df: pd.DataFrame, target: ec.Target, add_cols: dict) -> bool:
    res = False
    if df is not None:
        if not target or not target.type:
            logging.info("no target defined, printing first 100 rows:")
            print(df.head(100))
            res = True
        else:
            if target.type in (".csv"):
                res = push_csv(df, target)
            if target.type in (".xlsx"):
                res = push_excel(df, target)
            elif target.type_is_db:
                res = push_sql(df, target)
            elif target.type in ("pandas"):
                res = push_pandas(df, target)
            else:
                pass
    return res


def push_pandas(source_df: pd.DataFrame, target: ec.Target) -> bool:
    if not target.table:
        raise Exception("invalid table")
    if target.table not in el.staged_frames.keys():
        raise Exception("table not found in staged frames")

    el.staged_frames[target.table] = pd.concat(
        [el.staged_frames[target.table], source_df], ignore_index=True
    )
    return True


def push_sql(source_df: pd.DataFrame, target: ec.Target) -> bool:
    if not target.db_connection_string:
        raise Exception("invalid db_connection_string")
    if not target.table:
        raise Exception("invalid to_sql")
    kwargs = {}
    if target.type in ("mssql") and len(ec.supported_available_odbc_drivers()):
        kwargs["fast_executemany"] = True
    with sa.create_engine(target.db_connection_string, **kwargs).connect() as sqeng:
        if target.to_sql:
            kwargs = target.to_sql.model_dump()
        else:
            kwargs = {}
        source_df.to_sql(
            target.table,
            sqeng,
            schema=target.dbschema,
            index=False,
            if_exists="append",
            chunksize=1000,
            **kwargs,
        )
        sqeng.connection.commit()
        return True


def push_csv(source_df: pd.DataFrame, target: ec.Target) -> bool:
    if not target.url:
        raise Exception("no file path")
    if not os.path.exists(os.path.isfile(target.url)):
        raise Exception("invalid file path")

    if target.to_csv:
        kwargs = target.to_csv.model_dump()
    else:
        kwargs = {}

    source_df.to_csv(target.url, index=False, mode="a", header=False, **kwargs)

    return True


def push_excel(source_df: pd.DataFrame, target: ec.Target) -> bool:
    if not target.url:
        raise Exception("missing url")

    xlIO = el.fetch_excel_io(target.url)
    xlIO.set_sheet_df(target.sheet_name, source_df, target.if_exists)

    return True


def pull_sql(frame: ec.Frame, nrows=None, **kwargs) -> pd.DataFrame:
    if "norws" in kwargs:
        kwargs.pop("norws")
    if not frame.db_connection_string:
        raise Exception("invalid db_connection_string")
    if not frame.sqn:
        raise Exception("invalid sqn")
    with sa.create_engine(frame.db_connection_string).connect() as sqeng:
        stmt = sa.select(sa.text("*")).select_from(sa.text(frame.sqn)).limit(nrows)
        df = pd.read_sql(stmt, con=sqeng, **kwargs)
    return df


def build_sql(df: pd.DataFrame, target: ec.Frame, add_cols: dict) -> bool:
    if not target.db_connection_string:
        raise Exception("invalid db_connection_string")
    if not target.sqn:
        raise Exception("invalid sqn")
    if not target.table:
        raise Exception("invalid table")

    with sa.create_engine(target.db_connection_string).connect() as sqeng:
        sqeng.execute(sa.text(f"drop table if exists {target.sqn}"))

        # Use the first row to create the table structure
        df.head(1).to_sql(target.table, sqeng, schema=target.dbschema, index=False)

        # Delete the temporary row from the table
        sqeng.execute(sa.text(f"DELETE FROM {target.sqn}"))

        if add_cols:
            for col_name, col_val in add_cols.items():
                if col_val == ec.DynamicColumnValue.ROW_INDEX.value:
                    # Add an identity column to the table
                    sqeng.execute(
                        sa.text(
                            (
                                f"ALTER TABLE {target.sqn} ADD {col_name}"
                                " int identity(1,1) PRIMARY KEY "
                            )
                        )
                    )

        sqeng.connection.commit()
    return True


def build_csv(df: pd.DataFrame, target: ec.Frame) -> bool:
    if not target.url:
        raise Exception("invalid file_path")

    # save header row to csv, overwriting if exists
    df.head(0).to_csv(target.url, index=False, mode="w")

    return True


def build_excel_frame(df: pd.DataFrame, target: ec.Target) -> bool:
    if target.preparation_action == "create_replace_file":
        replace_file = True
    else:
        replace_file = False

    xlIO = el.fetch_excel_io(target.url, replace=replace_file)
    xlIO.set_sheet_df(target.sheet_name, df, target.if_exists)

    return True


def build_target(df: pd.DataFrame, target: ec.Frame, add_cols: dict) -> bool:
    if target.type_is_db:
        res = build_sql(df, target, add_cols)
    elif target.type in (".csv"):
        create_directory_if_not_exists(target.url)
        res = build_csv(df, target)
    elif target.type in (".xlsx"):
        create_directory_if_not_exists(target.url)
        res = build_excel_frame(
            pd.DataFrame(columns=df.columns, index=None, data=None), target
        )
    elif target.type in ("pandas"):
        res = build_pandas_frame(df, target)
    else:
        raise Exception("invalid target type")
    return res


def build_pandas_frame(df: pd.DataFrame, target: ec.Frame) -> bool:
    if not target.table:
        raise Exception("invalid table")

    if (
        target.table in el.staged_frames.keys()
        and target.if_exists == ec.TargetIfExistsValue.FAIL
    ):
        raise Exception(f"table {target.table} already exists in staged frames")
    else:
        el.staged_frames[target.table] = df.head(0)

    return True


def create_directory_if_not_exists(file_path: str):
    directory = os.path.dirname(file_path)
    if directory and not os.path.exists(directory):
        os.makedirs(directory)


def truncate_target(target: ec.Target) -> bool:
    if target.type_is_db:
        res = truncate_sql(target)
    elif target.type in (".csv"):
        res = truncate_csv(target)
    elif target.type in (".xlsx"):
        res = truncate_excel(target)
    elif target.type in ("pandas"):
        res = truncate_pandas(target)
    else:
        raise Exception("invalid target type")
    return res


def truncate_pandas(target):
    df = el.staged_frames[target.table]
    df.drop(df.index, axis=0, inplace=True)
    return True


def truncate_csv(target: ec.Target) -> bool:
    if not target.url:
        raise Exception("no file path")
    if not os.path.exists(os.path.isfile(target.url)):
        raise Exception("invalid file path")

    # read the first row of the file
    with open(target.url, "r") as f:
        reader = csv.reader(f)
        first_row = next(reader)

    # write the first row back to the file
    with open(target.url, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(first_row)

    return True


def truncate_excel(target: ec.Target) -> bool:
    if not target.url:
        raise Exception("no file path")
    if not os.path.exists(os.path.isfile(target.url)):
        raise Exception("invalid file path")

    # read the first row of the file
    with open(target.url, "r") as f:
        reader = csv.reader(f)
        first_row = next(reader)

    # write the first row back to the file
    with open(target.url, "w", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(first_row)

    return True


def clear_excel_sheet_after_row(file_path, sheet_name, row_start):
    # Load the workbook and select the sheet
    wb = load_workbook(file_path)
    sheet = wb[sheet_name]

    # Iterate over the rows
    for i in range(row_start, sheet.max_row + 1):
        for j in range(1, sheet.max_column + 1):
            # Clear the cell
            sheet.cell(row=i, column=j).value = None

    # Save the workbook
    wb.save(file_path)


def truncate_sql(target: ec.Target) -> bool:
    if not target.db_connection_string:
        raise Exception("invalid db_connection_string")
    with sa.create_engine(target.db_connection_string).connect() as sqeng:
        sqeng.execute(sa.text(f"truncate table {target.sqn}"))
        sqeng.connection.commit()
    return True


def frames_consistent(config: ec.Config) -> bool:
    target, source, add_cols, transform = get_configs(config)
    # fileBytes = fetch_file_bytes(target.url)
    # if len(fileBytes.getbuffer().tobytes()) == 0:
    #     return True

    # if target and target.type in ("pandas"):
    #     return True

    ignore_cols = []
    if add_cols:
        for k, v in add_cols.items():
            if v == ec.DynamicColumnValue.ROW_INDEX.value:
                ignore_cols.append(k)

    source_df = pull_frame(source, 100, add_cols, transform)
    source_df = add_columns(source_df, add_cols)
    target_df = pull_frame(target, 100)
    return data_frames_consistent(source_df, target_df, ignore_cols)


def data_frames_consistent(
    df1: pd.DataFrame, df2: pd.DataFrame, ignore_cols: list = []
) -> bool:
    res = True
    ignore_cols_set = set(ignore_cols)
    # Compare the column names and types
    source_cols = set(df1.columns.tolist()) - ignore_cols_set
    target_cols = set(df2.columns.tolist()) - ignore_cols_set

    if source_cols != target_cols:
        in_source = source_cols - target_cols
        in_target = target_cols - source_cols
        if in_source:
            logging.info("source has more columns:" + str(in_source))
        if in_target:
            logging.info("target has more columns:" + str(in_target))
        res = False
    else:
        for col in source_cols:
            # if nulls are returned from sql and object type is set in df
            if df2[col].dtype != "object" and df1[col].dtype != df2[col].dtype:
                logging.info(
                    f"{col} has a different data type source "
                    f"{df1[col].dtype} target {df2[col].dtype}"
                )
                res = False

    return res  # Table exists and has the same field names and types


def get_sql_data_type(dtype):
    if dtype == "int64":
        return "INT"
    elif dtype == "float64":
        return "FLOAT"
    elif dtype == "bool":
        return "BIT"
    elif dtype == "object":
        return "VARCHAR(MAX)"
    elif dtype == "datetime64":
        return "DATETIME"
    else:
        return "VARCHAR(MAX)"


def text_range_to_list(text):
    result = []
    segments = text.split(",")
    for segment in segments:
        if "-" in segment:
            start, end = map(int, segment.split("-"))
            result.extend(range(start, end + 1))
        else:
            result.append(int(segment))
    return result


def clean_page_numbers(page_numbers):
    if isinstance(page_numbers, int):
        res = [page_numbers]
    if isinstance(page_numbers, str):
        res = text_range_to_list(page_numbers)
    else:
        res = page_numbers
    return sorted(res)


def pull_pdf(file, laparams, **kwargs) -> pd.DataFrame:
    def get_first_char_from_text_box(tb) -> LTChar:
        for line in tb:
            for char in line:
                return char

    lap = LAParams()
    if laparams:
        for k, v in laparams.items():
            lap.__setattr__(k, v)

    if "page_numbers" in kwargs:
        kwargs["page_numbers"] = clean_page_numbers(kwargs["page_numbers"])

    pm_pages = extract_pages(file, laparams=lap, **kwargs)

    dict_res = {
        "page_index": [],
        "y0": [],
        "y1": [],
        "x0": [],
        "x1": [],
        "height": [],
        "width": [],
        "font_name": [],
        "font_size": [],
        "font_color": [],
        "text": [],
    }

    for p in pm_pages:
        for e in p:
            if isinstance(e, LTTextBox):
                first_char = get_first_char_from_text_box(e)
                dict_res["page_index"].append(
                    kwargs["page_numbers"][p.pageid - 1]
                    if "page_numbers" in kwargs
                    else p.pageid
                )
                dict_res["x0"].append(e.x0)
                dict_res["x1"].append(e.x1)
                dict_res["y0"].append(e.y0)
                dict_res["y1"].append(e.y1)
                dict_res["height"].append(e.height)
                dict_res["width"].append(e.width)
                dict_res["font_name"].append(first_char.fontname)
                dict_res["font_size"].append(first_char.height)
                dict_res["font_color"].append(
                    str(first_char.graphicstate.ncolor)
                    if not isinstance(first_char.graphicstate.ncolor, tuple)
                    else str(first_char.graphicstate.ncolor)
                )
                dict_res["text"].append(e.get_text().replace("\n", " ").rstrip())

    return pd.DataFrame(dict_res)


def pull_csv(file, clean_last_column, **kwargs):
    df = pd.read_csv(file, **kwargs)
    # check if last column is unnamed
    if (
        clean_last_column
        and isinstance(df.columns[-1], str)
        and df.columns[-1].startswith("Unnamed")
    ):
        # check if the last column is all null
        if df[df.columns[-1]].isnull().all():
            # drop the last column
            df = df.drop(df.columns[-1], axis=1)
    return df


def pull_fwf(file, **kwargs):
    df = pd.read_fwf(file, **kwargs)
    return df


def pull_xml(file, **kwargs):
    df = pd.read_xml(file, **kwargs)
    return df


def get_source_kwargs(read_x, frame: ec.Source, nrows: Optional[int] = None):
    kwargs = {}
    if read_x:
        kwargs = read_x.model_dump(exclude_none=True)

    for k, v in kwargs.items():
        if v == "None":
            kwargs[k] = None

    root_kwargs = (
        "nrows",
        "dtype",
        "sheet_name",
        "names",
        "encoding",
        "low_memory",
        "sep",
    )
    for k in root_kwargs:
        if hasattr(frame, k) and getattr(frame, k):
            if k == "dtype":
                dtypes = getattr(frame, "dtype")
                kwargs["dtype"] = {k: v for k, v in dtypes.items() if v != "date"}
            else:
                kwargs[k] = getattr(frame, k)

    if nrows:
        kwargs["nrows"] = nrows

    return kwargs


def get_target_kwargs(to_x, frame: ec.Target, nrows: Optional[int] = None):
    kwargs = {}
    if to_x:
        kwargs = to_x.model_dump(exclude_none=True)

    root_kwargs = (
        "nrows",
        "dtype",
        "sheet_name",
        "names",
        "encoding",
        "low_memory",
        "sep",
    )
    for k in root_kwargs:
        if hasattr(frame, k) and getattr(frame, k):
            kwargs[k] = getattr(frame, k)
    if nrows:
        kwargs["nrows"] = nrows

    return kwargs


def pull_frame(
    frame: Union[ec.Source, ec.Target],
    nrows: Optional[int] = None,
    add_cols: dict = {},
    transform: Optional[ec.Transform] = None,
    # dtype=None,
) -> pd.DataFrame:
    # logging.info(f"pulling frame {frame.file_path_dynamic}")
    if frame.type_is_db:
        kwargs = get_source_kwargs(None, frame, nrows)
        df = pull_sql(frame, **kwargs)
    elif frame.type in (".csv", ".tsv"):
        if isinstance(frame, ec.Source):
            clean_last_column = True
            # kwargs = get_source_kwargs(frame.read_csv, nrows, dtype)
            kwargs = get_source_kwargs(frame.read_csv, frame, nrows)
            # print(kwargs)
            if frame.type == ".tsv":
                kwargs["sep"] = "\t"

        else:
            clean_last_column = False
            kwargs = {}
        if "sep" not in kwargs.keys():
            kwargs["sep"] = ","
        df = pull_csv(frame.url, clean_last_column, **kwargs)

        # read first 10 rows of csv file with python csv reader into a list of rows
        with open(frame.url, "r", encoding="utf-8-sig") as f:
            row_scan_max = 10
            # get row count and update line_number for each line read
            row_scan = sum(
                1 for line_number, row in enumerate(f, 1) if line_number <= row_scan_max
            )
            f.seek(0)
            # take min of row count and 10
            # row_scan = 2
            reader = csv.reader(f, delimiter=kwargs["sep"])
            rows_n = [next(reader) for _ in range(row_scan)]

        # loop the values in add_cols
        for k, v in add_cols.items():
            # check if the value is a DynamicCellValue
            if (
                v
                and isinstance(v, str)
                and v[1:].upper() in ec.DynamicCellValue.__members__.keys()
            ):
                row, col = v[1:].upper().strip("R").split("C")
                row = int(row)
                col = int(col)
                # if v == "_r1c1":
                # get the cell value corresponding to the rxcx
                add_cols[k] = rows_n[row][col]

    elif frame.type and frame.type in (".xlsx", ".xls", ".xlsm", ".xlsb"):
        if isinstance(frame, ec.Source):
            # kwargs = get_source_kwargs(frame.read_excel, nrows, dtype)
            kwargs = get_source_kwargs(frame.read_excel, frame, nrows)

        elif isinstance(frame, ec.Target):
            kwargs = get_target_kwargs(frame.to_excel, frame, nrows)
        else:
            kwargs = {}
        xlIO = el.fetch_excel_io(frame.url)

        # loop the values in add_cols
        for k, v in add_cols.items():
            # check if the value is a DynamicCellValue
            if (
                v
                and isinstance(v, str)
                and v[1:].upper() in ec.DynamicCellValue.__members__.keys()
            ):
                row, col = v[1:].upper().strip("R").split("C")
                row = int(row)
                col = int(col)
                # get the cell value corresponding to the row/col
                add_cols[k] = xl.get_sheet_row(xlIO.fileIO, frame.sheet_name, row)[col]

        df = xlIO.pull_sheet(**kwargs)
    elif frame.type == ".fwf":
        if isinstance(frame, ec.Source):
            kwargs = get_source_kwargs(frame.read_fwf, frame, nrows)
        else:
            kwargs = {}

        df = pull_fwf(frame.url, **kwargs)
    elif frame.type == ".pdf":
        if isinstance(frame, ec.Source) and frame.extract_pages_pdf:
            kwargs = frame.extract_pages_pdf.model_dump(exclude_none=True)
            laparams = None
            if "laparams" in kwargs:
                laparams = kwargs.pop("laparams")
        else:
            kwargs = {}
        df = pull_pdf(frame.url, laparams=laparams, **kwargs)
    elif frame.type == ".xml":
        if isinstance(frame, ec.Source):
            kwargs = get_source_kwargs(frame.read_xml, frame)
        else:
            kwargs = {}
        if "nrows" in kwargs:
            kwargs.pop("nrows")
        df = pull_xml(frame.url, **kwargs)
        if nrows:
            df = df.head(nrows)
    elif frame.type in ("pandas"):
        if frame.table in el.staged_frames.keys():
            df = el.staged_frames[frame.table]
        else:
            raise Exception("pandas frame not found")
    else:
        raise Exception("unable to build df")
    if isinstance(df.columns, pd.MultiIndex):
        if transform and transform.stack:
            df = stack_columns(df, transform.stack)
        else:
            df = multiindex_to_singleindex(df)
    if isinstance(frame, ec.Source) and frame.filter:
        df = df.query(frame.filter)
    if transform and transform.melt:
        df = pd.melt(
            df,
            id_vars=transform.melt.id_vars,
            value_vars=transform.melt.value_vars,
            value_name=transform.melt.value_name,
            var_name=transform.melt.var_name,
        )
    if transform and transform.prql:
        with io.open(transform.prql) as file:
            prql = file.read()
        prqlo = prqlc.CompileOptions(target="sql.duckdb")
        dsql = prqlc.compile(prql, options=prqlo)
        df = duckdb.sql(dsql).df()
    if transform and transform.pivot:
        df = df.pivot(
            columns=transform.pivot.columns,
            values=transform.pivot.values,
            index=transform.pivot.index,
        )
        df.columns.name = None
        df.index.name = None
    if transform and transform.astype:
        df = df.astype(transform.astype.dtype)
    if hasattr(frame, "dtype") and frame.dtype:
        for k, v in frame.dtype.items():
            if v == "date" and not isinstance(type(df[k]), np.dtypes.DateTime64DType):
                df[k] = pd.to_datetime(df[k])
    return pd.DataFrame(df)


def stack_columns(df, stack: ec.Stack):
    # Define the primary column headers based on the first four columns
    primary_headers = list(df.columns[: stack.fixed_columns])

    # Extract the top-level column names from the primary headers
    top_level_headers, _ = zip(*primary_headers)

    # Set the DataFrame's index to the primary headers
    df = df.set_index(primary_headers)

    # Get the names of the newly set indices
    current_index_names = list(df.index.names[: stack.fixed_columns])

    # Create a dictionary to map the current index names to the top-level headers
    index_name_mapping = dict(zip(current_index_names, top_level_headers))

    # Rename the indices using the created mapping
    df.index.rename(index_name_mapping, inplace=True)

    # Stack the DataFrame based on the top-level columns
    df = df.stack(level=stack.stack_header)

    # Rename the new index created by the stacking operation
    df.index.rename({None: stack.stack_name}, inplace=True)

    # Reset the index for the resulting DataFrame
    df.reset_index(inplace=True)

    return df


def multiindex_to_singleindex(df, separator="_"):
    df.columns = [separator.join(map(str, col)).strip() for col in df.columns.values]
    return df


def get_configs(config):
    target = config.target
    source = config.source
    add_cols = config.add_cols.model_dump()
    transform = config.transform

    return target, source, add_cols, transform


def add_columns(df: pd.DataFrame, add_cols: dict) -> pd.DataFrame:
    if add_cols:
        for k, v in add_cols.items():
            if (
                k != "additionalProperties"
                and v != ec.DynamicColumnValue.ROW_INDEX.value
            ):
                df[k] = v
    return df


def ingest(config: ec.Config) -> bool:
    target, source, add_cols, transform = get_configs(config)
    consistent = frames_consistent(config)
    if (
        not target
        or not target.table
        or consistent
        or target.consistency == ec.TargetConsistencyValue.IGNORE.value
    ):
        source_df = pull_frame(source, config.nrows, add_cols, transform)
        source_df = add_columns(source_df, add_cols)
        return push_frame(source_df, target, add_cols)
    else:
        logging.error(target.table + ": Inconsistent, not saved.")
        return False


def build(config: ec.Config) -> bool:
    target, source, add_cols, transform = get_configs(config)
    if target and target.preparation_action != "no_action":
        action = target.preparation_action
        if action in ("create_replace", "create_replace_file"):
            # TODO, use caching to avoid pulling the same data twice
            df = pull_frame(source, 100, add_cols, transform)
            df = add_columns(df, add_cols)
            res = build_target(df, target, add_cols)
        elif action == "truncate":
            res = truncate_target(target)
        elif action == "fail":
            logging.error("Table Exists, failing")
            res = False
        else:
            res = True
    else:
        res = True
    return res


def detect(config: ec.Config) -> bool:
    _, source, _, _ = get_configs(config)
    source = source.model_copy()
    source.nrows = 100

    df = pull_frame(source)
    print(df.dtypes.to_dict())
    return True
